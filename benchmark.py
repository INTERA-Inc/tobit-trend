#!/usr/bin/env python3
"""
Acceptance test: compare Python TTA validation table against legacy R output.

Usage
-----
    python benchmark.py [control_file]

control_file defaults to 'benchmark_control.toml' in the current directory.

Control file format (TOML):
    python_csv          = "outputs/TTA_Validation_run.csv"
    r_csv               = "reference/TTA_Validation_R.csv"
    selected_wells      = []              # leave empty → interactive selection
    pass_threshold_pct  = 0.1            # relative error threshold (%)
    excel_output        = "benchmark_results.xlsx"  # set to "" to disable
"""
from __future__ import annotations

import sys
import tomllib
from datetime import date, datetime
from pathlib import Path
from typing import Optional

import numpy as np
import pandas as pd


# ---------------------------------------------------------------------------
# Column classification
# ---------------------------------------------------------------------------

NUMERIC_COLS: list[str] = [
    "Mean",
    "UCL",
    "LCL",
    "RiverStage_Coeff",
    "RiverStage_Coeff_stderror",
    "Date_Coeff",
    "Date_Coeff_stderror",
    "STD",
    "Intercept",
    "Intercept_stderror",
    "Lag",
    "XCoords",
    "YCoords",
    "Distance",
]

INTEGER_COLS: list[str] = ["N"]

BOOL_COLS: set[str] = {"Use_RiverStage"}
# removed: "Mean_Exceed", "UCL_Exceed"

PCT_STR_COLS: set[str] = {"PND"}

CATEGORICAL_COLS: list[str] = [
    "OU", "System", "Model", "Trend",
    "Transformation", "Trend_Break", "LAG_ORIGIN", "Analyte",
]
DEFAULT_CONTROL = "benchmark_control.toml"


# ---------------------------------------------------------------------------
# Value normalisation
# ---------------------------------------------------------------------------

def _to_numeric(val) -> Optional[float]:
    if val is None:
        return None
    s = str(val).strip()
    if s.lower() in ("nan", "nat", "none", ""):
        return None
    try:
        v = float(s)
        return None if np.isnan(v) else v
    except ValueError:
        return None


def _to_bool(val) -> Optional[bool]:
    if val is None:
        return None
    s = str(val).strip().lower()
    if s in ("nan", "none", ""):
        return None
    if s in ("true", "1", "yes"):
        return True
    if s in ("false", "0", "no"):
        return False
    return None


def _to_pct_int(val) -> Optional[int]:
    if val is None:
        return None
    s = str(val).strip().rstrip("%").strip()
    if s.lower() in ("nan", "none", ""):
        return None
    try:
        return round(float(s))
    except ValueError:
        return None


def _to_str(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    return None if s.lower() in ("nan", "nat", "none", "") else s


_TREND_BREAK_FMTS: tuple[str, ...] = ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y")


def _normalise_trend_break(val) -> Optional[str]:
    s = _to_str(val)
    if s is None:
        return None
    for fmt in _TREND_BREAK_FMTS:
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return s


def _normalise_trend(val) -> Optional[str]:
    s = _to_str(val)
    if s is None:
        return None
    if s.strip().lower() == "not performed":
        return None
    return s


# Analyte label variants produced by Python vs R for the same analyte.
_ANALYTE_CANONICAL: dict[str, str] = {
    "Hex. & Filt. Cr":       "Hex. & Filt. Chromium",
    "Hex. & Filt. Chromium": "Hex. & Filt. Chromium",
}


def _normalise_analyte(val) -> Optional[str]:
    s = _to_str(val)
    if s is None:
        return None
    return _ANALYTE_CANONICAL.get(s, s)


# Maps both Python and R raw LAG_ORIGIN strings to a shared canonical label.
_LAG_ORIGIN_CANONICAL: dict[str, str] = {
    "calculated":      "CrossCor",   # Python → R label
    "ULAG":            "WL",
    "no river stage":  "NoRS",
    "CrossCor":        "CrossCor",   # R labels (pass-through)
    "WL":              "WL",
    "NoRS":            "NoRS",
}


def _normalise_lag_origin(val) -> Optional[str]:
    s = _to_str(val)
    if s is None:
        return None
    return _LAG_ORIGIN_CANONICAL.get(s, s)


# ---------------------------------------------------------------------------
# Per-cell comparison
# ---------------------------------------------------------------------------

def _compare_cell(col: str, py_raw, r_raw, threshold_pct: float) -> dict:
    rec: dict = {"column": col, "python": py_raw, "r": r_raw}

    if col in NUMERIC_COLS or col in INTEGER_COLS:
        py = _to_numeric(py_raw)
        r  = _to_numeric(r_raw)
        both_na = py is None and r is None
        if both_na:
            rec.update(diff="—", kind="numeric", passed=True, both_na=True)
        elif py is None or r is None:
            rec.update(diff="one-NA", kind="numeric", passed=False, both_na=False)
        elif r == 0.0:
            passed = py == 0.0
            rec.update(diff="∞" if not passed else "0.00000%", kind="numeric",
                       passed=passed, both_na=False)
        else:
            rel = abs(py - r) / abs(r) * 100.0
            rec.update(diff=f"{rel:.5f}%", kind="numeric",
                       passed=rel <= threshold_pct, both_na=False)

    elif col in BOOL_COLS:
        py = _to_bool(py_raw)
        r  = _to_bool(r_raw)
        both_na = py is None and r is None
        rec.update(diff="—", kind="bool", passed=both_na or (py == r), both_na=both_na)

    elif col in PCT_STR_COLS:
        py = _to_pct_int(py_raw)
        r  = _to_pct_int(r_raw)
        both_na = py is None and r is None
        if both_na:
            rec.update(diff="—", kind="pct", passed=True, both_na=True)
        elif py is None or r is None:
            rec.update(diff="one-NA", kind="pct", passed=False, both_na=False)
        else:
            rec.update(diff=f"{abs(py - r)} pp", kind="pct", passed=(py == r), both_na=False)

    else:
        if col == "Trend_Break":
            py_s = _normalise_trend_break(py_raw)
            r_s  = _normalise_trend_break(r_raw)
        elif col == "LAG_ORIGIN":
            py_s = _normalise_lag_origin(py_raw)
            r_s  = _normalise_lag_origin(r_raw)
        elif col == "Trend":
            py_s = _normalise_trend(py_raw)
            r_s  = _normalise_trend(r_raw)
        elif col == "Analyte":
            py_s = _normalise_analyte(py_raw)
            r_s  = _normalise_analyte(r_raw)
        else:
            py_s = _to_str(py_raw)
            r_s  = _to_str(r_raw)
        both_na = py_s is None and r_s is None
        rec.update(diff="—", kind="categorical", passed=both_na or (py_s == r_s), both_na=both_na)

    return rec


# ---------------------------------------------------------------------------
# Well comparison
# ---------------------------------------------------------------------------

_ALL_COLS: list[str] = (
    NUMERIC_COLS
    + INTEGER_COLS
    + sorted(BOOL_COLS)
    + sorted(PCT_STR_COLS)
    + CATEGORICAL_COLS
)


def _compare_well(
    well: str,
    py_row: pd.Series,
    r_row: pd.Series,
    threshold_pct: float,
) -> list[dict]:
    results: list[dict] = []
    for col in _ALL_COLS:
        rec = _compare_cell(col, py_row.get(col, np.nan), r_row.get(col, np.nan), threshold_pct)
        rec["well"] = well
        results.append(rec)

    # When Trend is not performed / NA in both tools, the associated Lag value is
    # meaningless — Python carries the lag forward while R emits NaN. Override to pass.
    trend_rec = next((r for r in results if r["column"] == "Trend"), None)
    lag_rec   = next((r for r in results if r["column"] == "Lag"),   None)
    if trend_rec is not None and lag_rec is not None and trend_rec.get("both_na"):
        lag_rec.update(diff="—", passed=True, both_na=True)

    return results


# ---------------------------------------------------------------------------
# Formatting helper (used by Excel writer)
# ---------------------------------------------------------------------------

def _fmt(val) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return "NaN"
    if isinstance(val, float):
        return f"{val:.6g}"
    return str(val)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _write_excel_report(
    all_results: list[dict],
    selected: list[str],
    threshold_pct: float,
    py_csv: Path,
    r_csv: Path,
    out_path: Path,
    overall_passed: bool,
) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl not installed; cannot write Excel output. pip install openpyxl",
              file=sys.stderr)
        return

    # Shared styles
    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")   # light green
    FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")   # light red
    NA_FILL   = PatternFill("solid", fgColor="EDEDED")   # light grey
    HDR_FILL  = PatternFill("solid", fgColor="2F5496")   # dark blue
    META_FILL = PatternFill("solid", fgColor="D9E1F2")   # pale blue
    PASS_BIG  = PatternFill("solid", fgColor="375623")   # dark green
    FAIL_BIG  = PatternFill("solid", fgColor="9C0006")   # dark red

    WHITE_BOLD  = Font(bold=True, color="FFFFFF")
    BOLD        = Font(bold=True)
    RESULT_FONT = Font(bold=True, size=16, color="FFFFFF")
    CENTER      = Alignment(horizontal="center", vertical="center")
    LEFT        = Alignment(horizontal="left",   vertical="center")

    failures = [r for r in all_results if not r["passed"]]
    total    = len(all_results)
    n_passed = sum(1 for r in all_results if r["passed"])
    n_failed = total - n_passed

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 64

    def _cell(row, col, value=None, font=None, fill=None, alignment=None):
        c = ws.cell(row, col, value)
        if font:      c.font      = font
        if fill:      c.fill      = fill
        if alignment: c.alignment = alignment
        return c

    r = 1
    _cell(r, 1, "TTA Benchmark Results", font=Font(bold=True, size=16))
    r += 1
    _cell(r, 1, f"Generated: {date.today()}", font=Font(italic=True, color="595959"))
    r += 2

    for label, value in [
        ("Python CSV", str(py_csv)),
        ("R CSV",      str(r_csv)),
        ("Threshold",  f"≤ {threshold_pct}%  relative error (numeric columns)"),
        ("Wells",      ", ".join(selected)),
    ]:
        _cell(r, 1, label, font=BOLD, fill=META_FILL, alignment=LEFT)
        _cell(r, 2, value, fill=META_FILL, alignment=LEFT)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1
    ws.row_dimensions[r].height = 30
    _cell(r, 1, "OVERALL RESULT",
          font=RESULT_FONT, fill=PASS_BIG if overall_passed else FAIL_BIG, alignment=CENTER)
    _cell(r, 2, "PASS" if overall_passed else "FAIL",
          font=RESULT_FONT, fill=PASS_BIG if overall_passed else FAIL_BIG, alignment=CENTER)
    r += 2

    for label, value in [
        ("Wells tested", len(selected)),
        ("Comparisons",  total),
        ("Passed",       n_passed),
        ("Failed",       n_failed),
    ]:
        _cell(r, 1, label, font=BOLD)
        _cell(r, 2, value)
        r += 1

    if failures:
        r += 1
        f_hdrs = ["Well", "Column", "Type", "Python", "R", "Diff", "Result"]
        for col_idx, h in enumerate(f_hdrs, 1):
            c = ws.cell(r, col_idx, h)
            c.font      = WHITE_BOLD
            c.fill      = HDR_FILL
            c.alignment = CENTER
        for col_letter, width in zip("ABCDEFG", [28, 34, 12, 20, 20, 14, 10]):
            ws.column_dimensions[col_letter].width = width
        r += 1
        for rec in failures:
            for col_idx, val in enumerate(
                [rec["well"], rec["column"], rec["kind"],
                 _fmt(rec["python"]), _fmt(rec["r"]), rec["diff"], "FAIL"], 1
            ):
                ws.cell(r, col_idx, val).fill = FAIL_FILL
            r += 1

    # ── Sheet 2: Detail ─────────────────────────────────────────────────────
    wd = wb.create_sheet("Detail")
    wd.sheet_view.showGridLines = False
    wd.freeze_panes = "A2"

    for col_letter, width in zip("ABCDEFG", [28, 36, 14, 20, 20, 14, 10]):
        wd.column_dimensions[col_letter].width = width

    for col_idx, h in enumerate(["Well", "Column", "Type", "Python", "R", "Diff", "Result"], 1):
        c = wd.cell(1, col_idx, h)
        c.font      = WHITE_BOLD
        c.fill      = HDR_FILL
        c.alignment = CENTER

    wd.auto_filter.ref = f"A1:G{len(all_results) + 1}"

    for row_idx, rec in enumerate(all_results, 2):
        fill = NA_FILL if rec.get("both_na") else (PASS_FILL if rec["passed"] else FAIL_FILL)
        for col_idx, val in enumerate(
            [rec["well"], rec["column"], rec["kind"],
             _fmt(rec["python"]), _fmt(rec["r"]), rec["diff"],
             "PASS" if rec["passed"] else "FAIL"], 1
        ):
            wd.cell(row_idx, col_idx, val).fill = fill

    wb.save(out_path)
    print(f"Excel report written to: {out_path}")


# ---------------------------------------------------------------------------
# Interactive well selection
# ---------------------------------------------------------------------------

def _select_wells_interactively(available: list[str]) -> list[str]:
    print("\nAvailable wells (present in both Python and R outputs):\n")
    col_w = max(len(w) for w in available)
    for i, name in enumerate(available, 1):
        print(f"  [{i:>4}] {name:<{col_w}}")
    print()
    print("Enter well numbers (comma/space separated), a name prefix, or 'all':")
    print("  Examples:  1,3,5   /   all   /   199-D")

    while True:
        try:
            raw = input("  > ").strip()
        except (EOFError, KeyboardInterrupt):
            sys.exit(0)

        if not raw:
            continue
        if raw.lower() == "all":
            return list(available)

        parts = [p for p in raw.replace(",", " ").split() if p]
        selected: list[str] = []
        bad: list[str] = []
        for p in parts:
            if p.isdigit():
                idx = int(p) - 1
                if 0 <= idx < len(available):
                    selected.append(available[idx])
                else:
                    bad.append(p)
            else:
                matches = [w for w in available if w.startswith(p)]
                selected.extend(matches) if matches else bad.append(p)

        if bad:
            print(f"  Unrecognised: {', '.join(bad)}. Try again.")
            continue
        if not selected:
            print("  No wells selected. Try again.")
            continue

        seen: set[str] = set()
        return [w for w in selected if not (w in seen or seen.add(w))]  # type: ignore[func-returns-value]


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    control_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_CONTROL)

    if not control_path.exists():
        print(f"ERROR: Control file not found: {control_path}", file=sys.stderr)
        print(
            f"       Copy benchmark_control.toml.example to {DEFAULT_CONTROL} "
            "and fill in the paths.",
            file=sys.stderr,
        )
        return 2

    with open(control_path, "rb") as fh:
        ctrl = tomllib.load(fh)

    py_csv            = Path(ctrl["python_csv"])
    r_csv             = Path(ctrl["r_csv"])
    threshold_pct     = float(ctrl.get("pass_threshold_pct", 0.1))
    configured_wells  = ctrl.get("selected_wells", [])
    excel_out_str     = ctrl.get("excel_output", "benchmark_results.xlsx")

    for path, label in [(py_csv, "Python"), (r_csv, "R")]:
        if not path.exists():
            print(f"ERROR: {label} CSV not found: {path}", file=sys.stderr)
            return 2

    df_py = pd.read_csv(py_csv, dtype=str)
    df_r  = pd.read_csv(r_csv,  dtype=str)

    for label, df in [("Python", df_py), ("R", df_r)]:
        if "Name" not in df.columns:
            print(f"ERROR: {label} CSV is missing the 'Name' column.", file=sys.stderr)
            return 2

    df_py["Name"] = df_py["Name"].str.strip()
    df_r["Name"]  = df_r["Name"].str.strip()

    py_wells = set(df_py["Name"].dropna())
    r_wells  = set(df_r["Name"].dropna())
    common   = sorted(py_wells & r_wells)

    if not common:
        print("ERROR: No wells in common between the two CSV files.", file=sys.stderr)
        return 2

    if configured_wells:
        selected = [w for w in configured_wells if w in common]
        if not selected:
            print("ERROR: None of the configured wells appear in both outputs.", file=sys.stderr)
            return 2
    else:
        selected = _select_wells_interactively(common)

    py_idx = df_py.set_index("Name")
    r_idx  = df_r.set_index("Name")

    all_results: list[dict] = []
    for well in selected:
        py_row = py_idx.loc[well]
        r_row  = r_idx.loc[well]
        if isinstance(py_row, pd.DataFrame):
            py_row = py_row.iloc[0]
        if isinstance(r_row, pd.DataFrame):
            r_row = r_row.iloc[0]
        all_results.extend(_compare_well(well, py_row, r_row, threshold_pct))

    overall_passed = all(r["passed"] for r in all_results)

    if excel_out_str:
        _write_excel_report(
            all_results=all_results,
            selected=selected,
            threshold_pct=threshold_pct,
            py_csv=py_csv,
            r_csv=r_csv,
            out_path=Path(excel_out_str),
            overall_passed=overall_passed,
        )

    return 0 if overall_passed else 1


if __name__ == "__main__":
    sys.exit(main())
