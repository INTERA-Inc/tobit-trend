#!/usr/bin/env python3
"""
filter_benchmark.py

Validates that apply_well_filters() correctly implements the well_filter_cols /
well_filter_modes / well_filter_values settings from config.toml (lines 25-35).

Run this script after editing those config values to confirm filtering behaves
as expected before a full workflow run.

Usage
-----
    python filter_benchmark.py [config_path]

config_path defaults to "configs/config.toml".

Outputs
-------
    filter_benchmark_results.xlsx  -- colour-coded PASS/FAIL report

Tests covered
-------------
  Functional  : empty filters, include (single/multi), exclude (single/multi),
                combined include+exclude, whitespace stripping, case-insensitive
                mode strings, no-match include, case-sensitive value matching.
  Error paths : length mismatch, missing column, invalid mode, bare-string
                filter_values entry, empty column name, empty mode string.
  Config live : reads filter settings from config.toml and validates them
                against the synthetic well table.
"""
from __future__ import annotations

import sys
import tomllib
from datetime import date
from pathlib import Path
from typing import Any, Optional

import pandas as pd

OUTPUT_XLSX = "filter_benchmark_results.xlsx"
DEFAULT_CONFIG = "configs/config.toml"

# ---------------------------------------------------------------------------
# Synthetic well table shared by all functional tests
# ---------------------------------------------------------------------------

WELL_TABLE = pd.DataFrame(
    {
        "NAME": [
            "W-001", "W-002", "W-003", "W-004", "W-005",
            "W-006", "W-007", "W-008", "W-009", "W-010",
        ],
        "OU": [
            "100-KR-4",   "100-KR-4",   "100-HR-3-D", "100-HR-3-D",
            "100-HR-3-H", "100-HR-3-H", "100-BC-5",   "100-BC-5",
            "200-W-1",    "200-W-1",
        ],
        "STATUS": [
            "ACTIVE",           "DECOMMISSIONED-V",
            "ACTIVE",           "DRILLING CANCELLED",
            "ACTIVE",           "MONITORING",
            "DECOMMISSIONED-V", "ACTIVE",
            "DRILLING CANCELLED", "ACTIVE",
        ],
    }
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _run_filter(
    well: pd.DataFrame,
    cols: list,
    modes: list,
    values: list,
) -> pd.DataFrame:
    from tta.utils import apply_well_filters
    return apply_well_filters(well=well, filter_cols=cols, filter_modes=modes, filter_values=values)


def _expect_raises(
    exc_type: type,
    cols: list,
    modes: list,
    values: list,
) -> Optional[str]:
    """Return None when exc_type is raised, or an error message string on failure."""
    from tta.utils import apply_well_filters
    try:
        apply_well_filters(
            well=WELL_TABLE.copy(), filter_cols=cols, filter_modes=modes, filter_values=values
        )
        return f"Expected {exc_type.__name__} but no exception was raised."
    except exc_type:
        return None
    except Exception as exc:
        return f"Expected {exc_type.__name__} but got {type(exc).__name__}: {exc}"


def _pass(name: str, description: str, detail: str = "") -> dict:
    return {"name": name, "description": description, "status": "PASS", "detail": detail}


def _fail(name: str, description: str, detail: str) -> dict:
    return {"name": name, "description": description, "status": "FAIL", "detail": detail}


def _skip(name: str, description: str, reason: str) -> dict:
    return {"name": name, "description": description, "status": "SKIP", "detail": reason}


# ---------------------------------------------------------------------------
# Functional tests
# ---------------------------------------------------------------------------

def test_empty_filters() -> dict:
    name = "empty_filters"
    desc = "Empty filter_cols → all wells returned unchanged"
    try:
        out = _run_filter(WELL_TABLE.copy(), [], [], [])
        n_in, n_out = len(WELL_TABLE), len(out)
        if n_out == n_in:
            return _pass(name, desc, f"All {n_in} wells returned")
        return _fail(name, desc, f"Expected {n_in} wells, got {n_out}")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_include_ou_single() -> dict:
    name = "include_ou_single"
    desc = "Include single OU value → only wells with that OU retained"
    target = "100-KR-4"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["OU"], ["include"], [[target]])
        expected_n = int((WELL_TABLE["OU"] == target).sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        if not (out["OU"] == target).all():
            return _fail(name, desc, f"Output contains OU values other than {target!r}")
        return _pass(name, desc, f"{actual_n}/{len(WELL_TABLE)} wells retained (OU={target!r})")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_include_ou_multiple() -> dict:
    name = "include_ou_multiple"
    desc = "Include multiple OU values → all matching wells retained, others removed"
    targets = ["100-KR-4", "100-HR-3-D"]
    try:
        out = _run_filter(WELL_TABLE.copy(), ["OU"], ["include"], [targets])
        expected_n = int(WELL_TABLE["OU"].isin(targets).sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        if not out["OU"].isin(targets).all():
            bad = out.loc[~out["OU"].isin(targets), "OU"].unique().tolist()
            return _fail(name, desc, f"Unexpected OU values in output: {bad}")
        return _pass(name, desc, f"{actual_n}/{len(WELL_TABLE)} wells retained ({targets})")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_exclude_status_single() -> dict:
    name = "exclude_status_single"
    desc = "Exclude single STATUS value → matching rows removed"
    excl = "DECOMMISSIONED-V"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["STATUS"], ["exclude"], [[excl]])
        if (out["STATUS"] == excl).any():
            return _fail(name, desc, f"Output still contains STATUS={excl!r}")
        expected_n = int((WELL_TABLE["STATUS"] != excl).sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        return _pass(name, desc, f"{actual_n}/{len(WELL_TABLE)} wells retained after excluding {excl!r}")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_exclude_status_multiple() -> dict:
    name = "exclude_status_multiple"
    desc = "Exclude multiple STATUS values → all matching rows removed"
    excl = ["DECOMMISSIONED-V", "DRILLING CANCELLED"]
    try:
        out = _run_filter(WELL_TABLE.copy(), ["STATUS"], ["exclude"], [excl])
        bad_mask = out["STATUS"].isin(excl)
        if bad_mask.any():
            still_present = out.loc[bad_mask, "STATUS"].unique().tolist()
            return _fail(name, desc, f"Output still contains excluded statuses: {still_present}")
        expected_n = int((~WELL_TABLE["STATUS"].isin(excl)).sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        return _pass(name, desc, f"{actual_n}/{len(WELL_TABLE)} wells retained after excluding {excl}")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_include_and_exclude_combined() -> dict:
    name = "include_and_exclude_combined"
    desc = "Include OU list then exclude STATUS list → both filters applied in sequence"
    target_ous = ["100-KR-4", "100-HR-3-D", "100-HR-3-H"]
    excl_status = ["DECOMMISSIONED-V", "DRILLING CANCELLED"]
    try:
        out = _run_filter(
            WELL_TABLE.copy(),
            ["OU", "STATUS"],
            ["include", "exclude"],
            [target_ous, excl_status],
        )
        if not out["OU"].isin(target_ous).all():
            bad = out.loc[~out["OU"].isin(target_ous), "OU"].unique().tolist()
            return _fail(name, desc, f"Output contains wells outside target OU list: {bad}")
        if out["STATUS"].isin(excl_status).any():
            bad = out.loc[out["STATUS"].isin(excl_status), "STATUS"].unique().tolist()
            return _fail(name, desc, f"Output still contains excluded statuses: {bad}")
        expected_n = int(
            (WELL_TABLE["OU"].isin(target_ous) & ~WELL_TABLE["STATUS"].isin(excl_status)).sum()
        )
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        return _pass(name, desc, f"{actual_n}/{len(WELL_TABLE)} wells after include+exclude")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_whitespace_in_filter_values() -> dict:
    name = "whitespace_in_filter_values"
    desc = "Leading/trailing spaces in filter value list are stripped before matching"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["OU"], ["include"], [["  100-KR-4  "]])
        expected_n = int((WELL_TABLE["OU"] == "100-KR-4").sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(
                name, desc,
                f"Expected {expected_n} wells (whitespace stripped), got {actual_n}",
            )
        return _pass(name, desc, f"Whitespace stripped: {actual_n} wells matched")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_mode_case_insensitive() -> dict:
    name = "mode_case_insensitive"
    desc = "Mode string 'INCLUDE' (uppercase) is normalised to 'include'"
    target = "100-KR-4"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["OU"], ["INCLUDE"], [[target]])
        expected_n = int((WELL_TABLE["OU"] == target).sum())
        actual_n = len(out)
        if actual_n != expected_n:
            return _fail(name, desc, f"Expected {expected_n} wells, got {actual_n}")
        return _pass(name, desc, f"'INCLUDE' accepted; {actual_n} wells matched")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_include_no_match() -> dict:
    name = "include_no_match"
    desc = "Include filter that matches no rows → empty DataFrame, no exception"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["OU"], ["include"], [["NONEXISTENT-OU"]])
        if len(out) == 0:
            return _pass(name, desc, "Empty DataFrame returned as expected")
        return _fail(name, desc, f"Expected 0 wells, got {len(out)}")
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


def test_case_sensitivity_values() -> dict:
    name = "case_sensitivity_values"
    desc = "Filter values are case-sensitive: lowercase name does not match uppercase data"
    excl_lower = "decommissioned-v"
    try:
        out = _run_filter(WELL_TABLE.copy(), ["STATUS"], ["exclude"], [[excl_lower]])
        # lowercase should NOT match 'DECOMMISSIONED-V' → those rows must survive
        expected_survivors = int((WELL_TABLE["STATUS"] == "DECOMMISSIONED-V").sum())
        actual_survivors = int((out["STATUS"] == "DECOMMISSIONED-V").sum())
        if actual_survivors == expected_survivors:
            return _pass(
                name, desc,
                f"Case mismatch not excluded: {actual_survivors} 'DECOMMISSIONED-V' rows retained",
            )
        return _fail(
            name, desc,
            f"Expected {expected_survivors} 'DECOMMISSIONED-V' rows to remain, "
            f"got {actual_survivors} (case-insensitive match occurred)",
        )
    except Exception as exc:
        return _fail(name, desc, f"Unexpected {type(exc).__name__}: {exc}")


# ---------------------------------------------------------------------------
# Error-path tests
# ---------------------------------------------------------------------------

def test_error_length_mismatch() -> dict:
    name = "error_length_mismatch"
    desc = "Mismatched lengths of filter_cols/filter_modes/filter_values → ValueError"
    err = _expect_raises(
        ValueError,
        cols=["OU", "STATUS"],
        modes=["include"],                              # too short
        values=[["100-KR-4"], ["DECOMMISSIONED-V"]],
    )
    if err is None:
        return _pass(name, desc, "ValueError raised as expected")
    return _fail(name, desc, err)


def test_error_missing_column() -> dict:
    name = "error_missing_column"
    desc = "Filter references a column absent from the well table → KeyError"
    err = _expect_raises(
        KeyError,
        cols=["NONEXISTENT_COLUMN"],
        modes=["include"],
        values=[["some_value"]],
    )
    if err is None:
        return _pass(name, desc, "KeyError raised as expected")
    return _fail(name, desc, err)


def test_error_invalid_mode() -> dict:
    name = "error_invalid_mode"
    desc = "Unrecognised mode string → ValueError"
    err = _expect_raises(
        ValueError,
        cols=["OU"],
        modes=["contains"],           # neither 'include' nor 'exclude'
        values=[["100-KR-4"]],
    )
    if err is None:
        return _pass(name, desc, "ValueError raised as expected")
    return _fail(name, desc, err)


def test_error_bare_string_in_values() -> dict:
    name = "error_bare_string_in_values"
    desc = "Bare string instead of list as a filter_values entry → TypeError"
    err = _expect_raises(
        TypeError,
        cols=["OU"],
        modes=["include"],
        values=["100-KR-4"],          # should be [["100-KR-4"]]
    )
    if err is None:
        return _pass(name, desc, "TypeError raised as expected")
    return _fail(name, desc, err)


def test_error_empty_column_name() -> dict:
    name = "error_empty_column_name"
    desc = "Empty string in filter_cols → ValueError"
    err = _expect_raises(
        ValueError,
        cols=[""],
        modes=["include"],
        values=[["100-KR-4"]],
    )
    if err is None:
        return _pass(name, desc, "ValueError raised as expected")
    return _fail(name, desc, err)


def test_error_empty_mode_string() -> dict:
    name = "error_empty_mode_string"
    desc = "Empty string in filter_modes → ValueError"
    err = _expect_raises(
        ValueError,
        cols=["OU"],
        modes=[""],
        values=[["100-KR-4"]],
    )
    if err is None:
        return _pass(name, desc, "ValueError raised as expected")
    return _fail(name, desc, err)


# ---------------------------------------------------------------------------
# Live config test
# ---------------------------------------------------------------------------

def test_config_toml_filters(config_path: Optional[Path]) -> dict:
    """Load filter settings from config.toml and apply them to the synthetic well table."""
    name = "config_toml_filters"
    desc = (
        "Filter settings from config.toml applied to synthetic well table: "
        "include/exclude constraints verified on the output"
    )

    if config_path is None or not config_path.exists():
        return _skip(name, desc, f"Config not found: {config_path}")

    try:
        with open(config_path, "rb") as fh:
            raw = tomllib.load(fh)
        pc = raw.get("prep_chemistry", {})
        cols   = list(pc.get("well_filter_cols",   []))
        modes  = list(pc.get("well_filter_modes",  []))
        values = [list(v) for v in pc.get("well_filter_values", [])]
    except Exception as exc:
        return _fail(name, desc, f"Could not parse config.toml: {type(exc).__name__}: {exc}")

    if not cols:
        return _skip(name, desc, "well_filter_cols is empty in config.toml; nothing to test")

    # Check all configured columns are present in the synthetic table.
    missing_cols = [c for c in cols if c not in WELL_TABLE.columns]
    if missing_cols:
        return _skip(
            name, desc,
            f"Column(s) {missing_cols} are in well_filter_cols but not in the synthetic "
            "test table (NAME, OU, STATUS). Add them to WELL_TABLE in filter_benchmark.py "
            "to enable this test.",
        )

    try:
        out = _run_filter(WELL_TABLE.copy(), cols, modes, values)
    except Exception as exc:
        return _fail(name, desc, f"{type(exc).__name__}: {exc}")

    issues: list[str] = []
    for col, mode, vals in zip(cols, modes, values):
        vals_set = {str(v).strip() for v in vals}
        col_out = out[col].astype(str).str.strip()
        mode_norm = str(mode).strip().lower()
        if mode_norm == "include":
            n_bad = int((~col_out.isin(vals_set)).sum())
            if n_bad:
                issues.append(f"{n_bad} output rows have {col!r} outside include list")
        elif mode_norm == "exclude":
            n_bad = int(col_out.isin(vals_set).sum())
            if n_bad:
                issues.append(f"{n_bad} output rows still have excluded {col!r} values")

    summary = (
        f"Config: cols={cols}, modes={modes}, values={values} | "
        f"Input {len(WELL_TABLE)} wells → {len(out)} retained"
    )

    if issues:
        return _fail(name, desc, "; ".join(issues) + " | " + summary)
    return _pass(name, desc, summary)


# ---------------------------------------------------------------------------
# Run all tests
# ---------------------------------------------------------------------------

ALL_TESTS = [
    test_empty_filters,
    test_include_ou_single,
    test_include_ou_multiple,
    test_exclude_status_single,
    test_exclude_status_multiple,
    test_include_and_exclude_combined,
    test_whitespace_in_filter_values,
    test_mode_case_insensitive,
    test_include_no_match,
    test_case_sensitivity_values,
    test_error_length_mismatch,
    test_error_missing_column,
    test_error_invalid_mode,
    test_error_bare_string_in_values,
    test_error_empty_column_name,
    test_error_empty_mode_string,
]


def run_all_tests(config_path: Optional[Path]) -> list[dict]:
    results: list[dict] = []
    for fn in ALL_TESTS:
        results.append(fn())
    results.append(test_config_toml_filters(config_path))
    return results


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _write_excel(results: list[dict], config_path: Optional[Path], out_path: Path) -> None:
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(
            "openpyxl not installed; cannot write Excel output.\n"
            "  pip install openpyxl",
            file=sys.stderr,
        )
        return

    PASS_FILL = PatternFill("solid", fgColor="C6EFCE")
    FAIL_FILL = PatternFill("solid", fgColor="FFC7CE")
    SKIP_FILL = PatternFill("solid", fgColor="EDEDED")
    HDR_FILL  = PatternFill("solid", fgColor="2F5496")
    META_FILL = PatternFill("solid", fgColor="D9E1F2")
    PASS_BIG  = PatternFill("solid", fgColor="375623")
    FAIL_BIG  = PatternFill("solid", fgColor="9C0006")
    SKIP_BIG  = PatternFill("solid", fgColor="595959")

    WHITE_BOLD  = Font(bold=True, color="FFFFFF")
    BOLD        = Font(bold=True)
    BIG_FONT    = Font(bold=True, size=16, color="FFFFFF")
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

    n_pass = sum(1 for r in results if r["status"] == "PASS")
    n_fail = sum(1 for r in results if r["status"] == "FAIL")
    n_skip = sum(1 for r in results if r["status"] == "SKIP")
    total  = len(results)

    overall_status = "FAIL" if n_fail > 0 else "PASS"
    overall_fill   = FAIL_BIG if n_fail > 0 else PASS_BIG

    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 60

    def _cell(row, col, value=None, font=None, fill=None, alignment=None):
        c = ws.cell(row, col, value)
        if font:      c.font      = font
        if fill:      c.fill      = fill
        if alignment: c.alignment = alignment
        return c

    r = 1
    _cell(r, 1, "TTA Filter Benchmark Results", font=Font(bold=True, size=16))
    r += 1
    _cell(r, 1, f"Generated: {date.today()}", font=Font(italic=True, color="595959"))
    r += 2

    cfg_label = str(config_path) if config_path else "not provided"
    for label, value in [
        ("Config file", cfg_label),
        ("Synthetic wells", f"{len(WELL_TABLE)} rows (NAME, OU, STATUS)"),
    ]:
        _cell(r, 1, label, font=BOLD, fill=META_FILL, alignment=LEFT)
        _cell(r, 2, value,            fill=META_FILL, alignment=LEFT)
        ws.row_dimensions[r].height = 16
        r += 1

    r += 1
    ws.row_dimensions[r].height = 32
    _cell(r, 1, "OVERALL RESULT", font=BIG_FONT, fill=overall_fill, alignment=CENTER)
    _cell(r, 2, overall_status,   font=BIG_FONT, fill=overall_fill, alignment=CENTER)
    r += 2

    for label, value in [
        ("Total tests",  total),
        ("Passed",       n_pass),
        ("Failed",       n_fail),
        ("Skipped",      n_skip),
    ]:
        _cell(r, 1, label, font=BOLD)
        _cell(r, 2, value)
        r += 1

    # Failure list
    failures = [x for x in results if x["status"] == "FAIL"]
    if failures:
        r += 1
        for col_idx, h in enumerate(["Test Name", "Description", "Reason"], 1):
            c = ws.cell(r, col_idx, h)
            c.font = WHITE_BOLD
            c.fill = HDR_FILL
            c.alignment = CENTER
        ws.column_dimensions["A"].width = 36
        ws.column_dimensions["B"].width = 55
        ws.column_dimensions["C"].width = 60
        r += 1
        for rec in failures:
            ws.cell(r, 1, rec["name"]).fill        = FAIL_FILL
            ws.cell(r, 2, rec["description"]).fill  = FAIL_FILL
            ws.cell(r, 3, rec["detail"]).fill        = FAIL_FILL
            ws.row_dimensions[r].height = 30
            r += 1

    # ── Sheet 2: Detail ─────────────────────────────────────────────────────
    wd = wb.create_sheet("Detail")
    wd.sheet_view.showGridLines = False
    wd.freeze_panes = "A2"

    col_widths = {"A": 36, "B": 55, "C": 10, "D": 65}
    for col_letter, width in col_widths.items():
        wd.column_dimensions[col_letter].width = width

    headers = ["Test Name", "Description", "Status", "Detail"]
    for col_idx, h in enumerate(headers, 1):
        c = wd.cell(1, col_idx, h)
        c.font      = WHITE_BOLD
        c.fill      = HDR_FILL
        c.alignment = CENTER

    wd.auto_filter.ref = f"A1:D{len(results) + 1}"

    status_fill = {"PASS": PASS_FILL, "FAIL": FAIL_FILL, "SKIP": SKIP_FILL}

    for row_idx, rec in enumerate(results, 2):
        fill = status_fill.get(rec["status"], SKIP_FILL)
        for col_idx, val in enumerate(
            [rec["name"], rec["description"], rec["status"], rec["detail"]], 1
        ):
            c = wd.cell(row_idx, col_idx, val)
            c.fill      = fill
            c.alignment = LEFT
        wd.row_dimensions[row_idx].height = 28

    wb.save(out_path)
    print(f"Excel report written → {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    config_path: Optional[Path] = None

    if len(sys.argv) > 1:
        config_path = Path(sys.argv[1])
    else:
        default = Path(DEFAULT_CONFIG)
        if default.exists():
            config_path = default

    if config_path is not None and not config_path.exists():
        print(f"WARNING: config path not found: {config_path}", file=sys.stderr)
        config_path = None

    print("Running filter benchmark tests...")
    print(f"  Config: {config_path or '(none)'}")
    print(f"  Synthetic well table: {len(WELL_TABLE)} rows\n")

    results = run_all_tests(config_path)

    n_pass = sum(1 for r in results if r["status"] == "PASS")
    n_fail = sum(1 for r in results if r["status"] == "FAIL")
    n_skip = sum(1 for r in results if r["status"] == "SKIP")

    width = max(len(r["name"]) for r in results)
    for rec in results:
        marker = {"PASS": "✓", "FAIL": "✗", "SKIP": "-"}.get(rec["status"], "?")
        print(f"  [{marker}] {rec['name']:<{width}}  {rec['status']}")
        if rec["status"] in ("FAIL", "SKIP") and rec["detail"]:
            # indent detail under the test line
            for line in rec["detail"].split(" | "):
                print(f"        {line}")

    print(f"\n  Total: {len(results)}  Passed: {n_pass}  Failed: {n_fail}  Skipped: {n_skip}")

    _write_excel(results, config_path, Path(OUTPUT_XLSX))

    return 1 if n_fail > 0 else 0


if __name__ == "__main__":
    sys.exit(main())
