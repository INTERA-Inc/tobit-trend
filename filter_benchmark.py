#!/usr/bin/env python3
"""
filter_benchmark.py

Validates filter changes in config.toml without running the full TTA workflow.

Covers two independent filter groups:

  WELL FILTERS  (well_filter_cols / well_filter_modes / well_filter_values)
      Operates on the well info CSV.  Determines which wells enter the workflow
      at all.  Results are compared to a saved snapshot so the user can see
      exactly which wells moved in or out after a config change.

  CHEMISTRY FILTERS  (reviewq_remove_patterns / collection_purpose_exclude)
      Operates on the prepared chemistry parquet.  Determines which individual
      sample rows are kept.  Reports how many rows each filter removes and which
      wells are most affected.

Usage
-----
    python filter_benchmark.py [config_path]

config_path defaults to "configs/config.toml".

Outputs
-------
    filter_benchmark_results.xlsx   -- colour-coded Excel report
    filter_benchmark_snapshot.json  -- saved after each run for comparison
"""
from __future__ import annotations

import json
import sys
import tomllib
from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

OUTPUT_XLSX    = "filter_benchmark_results.xlsx"
SNAPSHOT_FILE  = "filter_benchmark_snapshot.json"
DEFAULT_CONFIG = "configs/config.toml"


# ---------------------------------------------------------------------------
# Config loading
# ---------------------------------------------------------------------------

def _load_well_config(raw: dict) -> tuple[Path, list, list, list]:
    """Return (well_file, cols, modes, values) from parsed TOML."""
    well_file = Path(raw["calculate_distance"]["well_info_well"])
    pc = raw["prep_chemistry"]
    cols   = list(pc.get("well_filter_cols",   []))
    modes  = list(pc.get("well_filter_modes",  []))
    values = [list(v) for v in pc.get("well_filter_values", [])]
    return well_file, cols, modes, values


def _load_chem_config(raw: dict) -> tuple[list[Path], list, list, float, str, pd.Timestamp, pd.Timestamp]:
    """Return chemistry filter settings from parsed TOML."""
    pc  = raw["prep_chemistry"]
    gs  = raw["global_settings"]
    chem_files        = [Path(p) for p in pc["raw_chemistry_files"]]
    reviewq_patterns  = list(pc.get("reviewq_remove_patterns",   []))
    purpose_exclude   = list(pc.get("collection_purpose_exclude", []))
    mdl_sub           = float(pc.get("mdl_sub_if_nonpositive_missing", 1.0))
    duplicate_handling = str(pc.get("duplicate_handling", "daily_avg"))
    min_date = pd.Timestamp(gs["global_min_date"])
    max_date = pd.Timestamp(
        pc.get("max_date", gs["global_max_date"])
    )
    return chem_files, reviewq_patterns, purpose_exclude, mdl_sub, duplicate_handling, min_date, max_date


# ---------------------------------------------------------------------------
# Well filtering (stepwise, with exclusion attribution)
# ---------------------------------------------------------------------------

def _apply_well_filters_stepwise(
    well: pd.DataFrame,
    cols: list,
    modes: list,
    values: list,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Apply well filters one at a time so each excluded well can be attributed
    to the specific filter that removed it.

    Returns (included, excluded) where excluded has an EXCLUSION_REASON column.
    """
    from tta.utils import apply_well_filters

    current = well.copy()
    attribution: dict[str, str] = {}

    for i, (col, mode, vals) in enumerate(zip(cols, modes, values), 1):
        before  = set(current["NAME"].astype(str))
        current = apply_well_filters(current, [col], [mode], [vals])
        after   = set(current["NAME"].astype(str))
        vals_display = ", ".join(str(v) for v in vals)
        reason = f"Filter {i} — {col} {mode}: [{vals_display}]"
        for name in before - after:
            attribution[name] = reason

    included      = current.copy()
    excluded_names = set(well["NAME"].astype(str)) - set(included["NAME"].astype(str))
    excluded       = well[well["NAME"].astype(str).isin(excluded_names)].copy()
    excluded["EXCLUSION_REASON"] = excluded["NAME"].astype(str).map(attribution)
    return included, excluded


# ---------------------------------------------------------------------------
# Chemistry filtering (stepwise, with per-well impact summary)
# ---------------------------------------------------------------------------

def _apply_chem_filters_stepwise(
    chem_files: list[Path],
    reviewq_patterns: list,
    purpose_exclude: list,
    min_date: pd.Timestamp,
    max_date: pd.Timestamp,
) -> tuple[dict, pd.DataFrame]:
    """
    Load chemistry data, apply filters in order, and return row-level stats.

    Filters applied (matching chemistry_import_01.py order):
      1. Date range  (global_min_date … global_max_date)
      2. REVIEWQ     (rows whose REVIEWQ contains any reviewq_remove_patterns value)
      3. COLLECTION_PURPOSE  (rows whose PURPOSE is in collection_purpose_exclude)

    Returns
    -------
    stats    : summary counts and per-filter breakdowns
    per_well : DataFrame — one row per well, columns:
               NAME | Total Rows | Removed (REVIEWQ) | Removed (PURPOSE) |
               Remaining | % Retained
    """
    from tta.preprocessing.chemistry_import_01 import read_prepared_chemistry

    parts = []
    for f in chem_files:
        if not f.exists():
            raise FileNotFoundError(f"Chemistry file not found: {f}")
        parts.append(read_prepared_chemistry(str(f)))

    chem = pd.concat(parts, ignore_index=True)
    chem["EVENT"] = pd.to_datetime(chem["EVENT"], errors="coerce")
    n_raw = len(chem)

    # 1. Date range
    chem_dated = chem.loc[
        (chem["EVENT"] >= min_date) & (chem["EVENT"] <= max_date)
    ].copy()
    n_after_date   = len(chem_dated)
    n_removed_date = n_raw - n_after_date

    # 2. REVIEWQ filter
    reviewq_breakdown: dict[str, int] = {}
    if "REVIEWQ" in chem_dated.columns and reviewq_patterns:
        reviewq_col  = chem_dated["REVIEWQ"].fillna("").astype(str)
        reviewq_mask = pd.Series(False, index=chem_dated.index)
        for pat in reviewq_patterns:
            matched = reviewq_col.str.contains(pat, regex=False)
            reviewq_breakdown[pat] = int(matched.sum())
            reviewq_mask |= matched
    else:
        reviewq_mask = pd.Series(False, index=chem_dated.index)

    n_removed_reviewq = int(reviewq_mask.sum())
    chem_after_reviewq = chem_dated.loc[~reviewq_mask].copy()

    # 3. COLLECTION_PURPOSE filter
    purpose_breakdown: dict[str, int] = {}
    if "COLLECTION_PURPOSE" in chem_after_reviewq.columns and purpose_exclude:
        purpose_col  = chem_after_reviewq["COLLECTION_PURPOSE"].astype(str)
        purpose_mask = purpose_col.isin(purpose_exclude)
        for code in purpose_exclude:
            count = int((purpose_col == code).sum())
            if count > 0:
                purpose_breakdown[code] = count
    else:
        purpose_mask = pd.Series(False, index=chem_after_reviewq.index)

    n_removed_purpose = int(purpose_mask.sum())
    chem_final = chem_after_reviewq.loc[~purpose_mask].copy()

    # Per-well impact summary
    pw_total    = chem_dated.groupby("NAME").size().rename("Total Rows")
    pw_reviewq  = (chem_dated.loc[reviewq_mask]
                   .groupby("NAME").size().rename("Removed (REVIEWQ)"))
    pw_purpose  = (chem_after_reviewq.loc[purpose_mask]
                   .groupby("NAME").size().rename("Removed (PURPOSE)"))
    pw_remaining = chem_final.groupby("NAME").size().rename("Remaining")

    per_well = (pd.concat([pw_total, pw_reviewq, pw_purpose, pw_remaining], axis=1)
                .fillna(0).astype(int).reset_index())
    per_well.columns = ["NAME", "Total Rows", "Removed (REVIEWQ)", "Removed (PURPOSE)", "Remaining"]
    per_well["% Retained"] = (
        per_well["Remaining"] / per_well["Total Rows"].replace(0, pd.NA) * 100
    ).round(1)
    per_well = per_well.sort_values(
        ["Removed (REVIEWQ)", "Removed (PURPOSE)"], ascending=False
    ).reset_index(drop=True)

    stats = {
        "n_raw":              n_raw,
        "n_after_date":       n_after_date,
        "n_removed_date":     n_removed_date,
        "date_range":         f"{min_date.date()} → {max_date.date()}",
        "n_removed_reviewq":  n_removed_reviewq,
        "reviewq_breakdown":  reviewq_breakdown,
        "n_removed_purpose":  n_removed_purpose,
        "purpose_breakdown":  purpose_breakdown,
        "n_final":            len(chem_final),
        "n_wells_affected":   int(((per_well["Removed (REVIEWQ)"] + per_well["Removed (PURPOSE)"]) > 0).sum()),
        "n_wells_emptied":    int((per_well["Remaining"] == 0).sum()),
    }

    return stats, per_well


# ---------------------------------------------------------------------------
# Snapshot
# ---------------------------------------------------------------------------

def _load_snapshot() -> Optional[dict]:
    p = Path(SNAPSHOT_FILE)
    if not p.exists():
        return None
    try:
        with open(p, encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return None


def _save_snapshot(
    well_file: Path,
    cols: list, modes: list, values: list,
    n_well_total: int,
    included_names: list[str],
    chem_stats: Optional[dict],
) -> None:
    snap: dict = {
        "run_timestamp":  datetime.now().isoformat(timespec="seconds"),
        "well_file":      str(well_file),
        "filter_cols":    cols,
        "filter_modes":   modes,
        "filter_values":  values,
        "n_well_total":   n_well_total,
        "n_included":     len(included_names),
        "included_wells": sorted(included_names),
    }
    if chem_stats:
        snap["chem"] = {
            "n_raw":             chem_stats["n_raw"],
            "n_after_date":      chem_stats["n_after_date"],
            "n_removed_reviewq": chem_stats["n_removed_reviewq"],
            "n_removed_purpose": chem_stats["n_removed_purpose"],
            "n_final":           chem_stats["n_final"],
        }
    with open(SNAPSHOT_FILE, "w", encoding="utf-8") as fh:
        json.dump(snap, fh, indent=2)


def _compute_well_diff(prev: dict, current_included: list[str]) -> tuple[list, list]:
    prev_set = set(prev.get("included_wells", []))
    curr_set = set(current_included)
    return sorted(curr_set - prev_set), sorted(prev_set - curr_set)


def _compute_chem_diff(prev: dict, chem_stats: dict) -> dict:
    prev_chem = prev.get("chem")
    if not prev_chem:
        return {}
    return {
        "d_removed_reviewq": chem_stats["n_removed_reviewq"] - prev_chem["n_removed_reviewq"],
        "d_removed_purpose": chem_stats["n_removed_purpose"] - prev_chem["n_removed_purpose"],
        "d_final":           chem_stats["n_final"]           - prev_chem["n_final"],
        "prev_n_final":      prev_chem["n_final"],
    }


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------

def _write_excel(
    config_path: Path,
    # well
    well_file: Path,
    cols: list, modes: list, values: list,
    included: pd.DataFrame,
    excluded: pd.DataFrame,
    prev_snap: Optional[dict],
    added_wells: list, removed_wells: list,
    # chemistry (None when chem files are unavailable)
    chem_files: list[Path],
    reviewq_patterns: list,
    purpose_exclude: list,
    mdl_sub: float,
    duplicate_handling: str,
    chem_stats: Optional[dict],
    per_well_chem: Optional[pd.DataFrame],
    chem_diff: dict,
    out_path: Path,
) -> None:
    if not HAS_OPENPYXL:
        print("openpyxl not installed — cannot write Excel.\n  pip install openpyxl",
              file=sys.stderr)
        return

    # Shared styles
    INCL_FILL  = PatternFill("solid", fgColor="C6EFCE")
    EXCL_FILL  = PatternFill("solid", fgColor="FFC7CE")
    ADDED_FILL = PatternFill("solid", fgColor="FFEB9C")
    HDR_FILL   = PatternFill("solid", fgColor="2F5496")
    META_FILL  = PatternFill("solid", fgColor="D9E1F2")
    GREY_FILL  = PatternFill("solid", fgColor="EDEDED")
    WARN_FILL  = PatternFill("solid", fgColor="FCE4D6")
    CHEM_FILL  = PatternFill("solid", fgColor="DAE8FC")

    WHITE_BOLD = Font(bold=True, color="FFFFFF")
    BOLD       = Font(bold=True)
    ITALIC     = Font(italic=True, color="595959")
    CENTER     = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT       = Alignment(horizontal="left",   vertical="center", wrap_text=False)

    n_total    = len(included) + len(excluded)
    n_included = len(included)
    n_excluded = len(excluded)

    wb = openpyxl.Workbook()

    # ── helpers ───────────────────────────────────────────────────────────────

    def _hdr(ws, row: int, headers: list, widths: list) -> None:
        for ci, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row, ci, h)
            c.font = WHITE_BOLD; c.fill = HDR_FILL; c.alignment = CENTER
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _row(ws, r: int, label: str, value, label_fill, value_fill=None) -> None:
        c1 = ws.cell(r, 1, label)
        c2 = ws.cell(r, 2, value)
        c1.fill = label_fill; c1.font = BOLD; c1.alignment = LEFT
        c2.fill = value_fill or label_fill; c2.alignment = LEFT

    def _section(ws, r: int, title: str) -> int:
        ws.cell(r, 1, title).font = BOLD
        return r + 1

    def _data_sheet(title: str, df: pd.DataFrame, display_cols: list, fill: PatternFill) -> None:
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        available = [c for c in display_cols if c in df.columns]
        if df.empty or not available:
            ws.cell(1, 1, "No data").font = ITALIC
            return
        widths = [max(len(c) + 4, 18) for c in available]
        for i, col in enumerate(available):
            if col == "EXCLUSION_REASON":
                widths[i] = 55
            if col == "% Retained":
                widths[i] = 14
        _hdr(ws, 1, available, widths)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(available))}{len(df)+1}"
        for ri, (_, row) in enumerate(df[available].iterrows(), 2):
            for ci, val in enumerate(row, 1):
                c = ws.cell(ri, ci, "" if pd.isna(val) else val)
                c.fill = fill; c.alignment = LEFT

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 55

    r = 1
    ws.cell(r, 1, "TTA Filter Benchmark Report").font = Font(bold=True, size=16)
    r += 1
    ws.cell(r, 1, f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}").font = ITALIC
    r += 2

    # Metadata
    for label, value in [("Config file", str(config_path)), ("Well file", str(well_file))]:
        _row(ws, r, label, value, META_FILL); r += 1
    r += 1

    # ── Well filter settings ──
    r = _section(ws, r, "Well filter settings")
    _hdr(ws, r, ["#", "Column", "Mode", "Values"], [6, 18, 12, 55]); r += 1
    if cols:
        for i, (col, mode, vals) in enumerate(zip(cols, modes, values), 1):
            fill = INCL_FILL if str(mode).strip().lower() == "include" else EXCL_FILL
            vals_str = ", ".join(str(v) for v in vals)
            for ci, val in enumerate([i, col, mode, vals_str], 1):
                ws.cell(r, ci, val).fill = fill
                ws.cell(r, ci).alignment = LEFT
            r += 1
    else:
        ws.cell(r, 1, "No well filters configured").font = ITALIC; r += 1
    r += 1

    # ── Well counts ──
    r = _section(ws, r, "Well counts")
    for label, value, fill in [
        ("Total wells in file",    n_total,    GREY_FILL),
        ("Included after filters", n_included, INCL_FILL),
        ("Excluded by filters",    n_excluded, EXCL_FILL),
    ]:
        _row(ws, r, label, value, fill); r += 1

    if not excluded.empty and "EXCLUSION_REASON" in excluded.columns:
        r += 1
        r = _section(ws, r, "Exclusions by filter")
        _hdr(ws, r, ["Filter", "Wells removed"], [55, 16]); r += 1
        for reason, count in excluded["EXCLUSION_REASON"].value_counts().items():
            _row(ws, r, reason, int(count), EXCL_FILL); r += 1
    r += 1

    # ── Well change summary vs previous run ──
    if prev_snap is not None:
        prev_ts   = prev_snap.get("run_timestamp", "unknown")
        prev_ninc = prev_snap.get("n_included", "?")
        r = _section(ws, r, "Well filter changes vs previous run")
        ws.cell(r, 1, f"Previous run: {prev_ts}  ({prev_ninc} wells included)").font = ITALIC
        r += 1
        for label, value, fill in [
            ("Wells newly included  (filter loosened)",  len(added_wells),   ADDED_FILL),
            ("Wells newly excluded  (filter tightened)", len(removed_wells), EXCL_FILL),
        ]:
            _row(ws, r, label, value, fill); r += 1
        if added_wells or removed_wells:
            ws.cell(r, 1, "→ See 'Well Changes' sheet for the full list").font = ITALIC; r += 1
    else:
        ws.cell(r, 1, "No previous snapshot — run again after editing the config to see a diff").font = ITALIC
        r += 1
    r += 1

    # ── Chemistry filter settings ──
    r = _section(ws, r, "Chemistry filter settings")
    _hdr(ws, r, ["Setting", "Values"], [38, 55]); r += 1

    setting_rows = [
        ("reviewq_remove_patterns",   ", ".join(reviewq_patterns)  or "(none)"),
        ("collection_purpose_exclude", ", ".join(purpose_exclude)   or "(none)"),
        ("mdl_sub_if_nonpositive_missing", str(mdl_sub)),
        ("duplicate_handling",         duplicate_handling),
    ]
    for label, value in setting_rows:
        ws.cell(r, 1, label).fill = CHEM_FILL
        ws.cell(r, 1).font = BOLD
        ws.cell(r, 1).alignment = LEFT
        ws.cell(r, 2, value).fill = CHEM_FILL
        ws.cell(r, 2).alignment = LEFT
        r += 1
    r += 1

    # ── Chemistry row counts ──
    if chem_stats:
        r = _section(ws, r, "Chemistry row counts")
        rows_data = [
            ("Rows in chemistry file",                  chem_stats["n_raw"],             GREY_FILL),
            (f"After date range ({chem_stats['date_range']})", chem_stats["n_after_date"], GREY_FILL),
            ("  Removed by date range",                 chem_stats["n_removed_date"],    WARN_FILL if chem_stats["n_removed_date"] else GREY_FILL),
            ("  Removed by reviewq_remove_patterns",    chem_stats["n_removed_reviewq"], EXCL_FILL if chem_stats["n_removed_reviewq"] else GREY_FILL),
            ("  Removed by collection_purpose_exclude", chem_stats["n_removed_purpose"], EXCL_FILL if chem_stats["n_removed_purpose"] else GREY_FILL),
            ("Remaining rows (pass all filters)",       chem_stats["n_final"],           INCL_FILL),
            ("Wells with any rows removed",             chem_stats["n_wells_affected"],  WARN_FILL if chem_stats["n_wells_affected"] else GREY_FILL),
            ("Wells completely emptied",                chem_stats["n_wells_emptied"],   EXCL_FILL if chem_stats["n_wells_emptied"] else GREY_FILL),
        ]
        for label, value, fill in rows_data:
            _row(ws, r, label, value, fill); r += 1
        r += 1

        # REVIEWQ breakdown
        if chem_stats["reviewq_breakdown"]:
            r = _section(ws, r, "REVIEWQ removals by pattern")
            _hdr(ws, r, ["Pattern", "Rows removed"], [20, 16]); r += 1
            for pat, count in chem_stats["reviewq_breakdown"].items():
                _row(ws, r, pat, count, EXCL_FILL if count else GREY_FILL); r += 1
            r += 1

        # COLLECTION_PURPOSE breakdown
        if chem_stats["purpose_breakdown"]:
            r = _section(ws, r, "COLLECTION_PURPOSE removals by code")
            _hdr(ws, r, ["Purpose code", "Rows removed"], [20, 16]); r += 1
            for code, count in sorted(chem_stats["purpose_breakdown"].items(),
                                      key=lambda x: -x[1]):
                _row(ws, r, code, count, EXCL_FILL); r += 1
            r += 1

        # Chemistry diff vs snapshot
        if chem_diff:
            prev_ts = prev_snap.get("run_timestamp", "unknown")  # type: ignore[union-attr]
            r = _section(ws, r, "Chemistry filter changes vs previous run")
            ws.cell(r, 1, f"Previous run: {prev_ts}  ({chem_diff['prev_n_final']:,} rows remaining)").font = ITALIC
            r += 1
            for label, delta, base_fill in [
                ("Change in rows removed by REVIEWQ",  chem_diff["d_removed_reviewq"], EXCL_FILL),
                ("Change in rows removed by PURPOSE",  chem_diff["d_removed_purpose"], EXCL_FILL),
                ("Change in rows remaining",           chem_diff["d_final"],           INCL_FILL),
            ]:
                prefix = "+" if delta > 0 else ""
                fill   = base_fill if delta != 0 else GREY_FILL
                _row(ws, r, label, f"{prefix}{delta:,}", fill); r += 1
    else:
        ws.cell(r, 1, "Chemistry files not found — chemistry filter validation skipped").font = ITALIC
        r += 1

    # ── Sheet 2: Included wells ───────────────────────────────────────────────
    _data_sheet("Well — Included", included, ["NAME"] + cols, INCL_FILL)

    # ── Sheet 3: Excluded wells ───────────────────────────────────────────────
    _data_sheet("Well — Excluded", excluded, ["NAME"] + cols + ["EXCLUSION_REASON"], EXCL_FILL)

    # ── Sheet 4: Well Changes (only when something changed) ───────────────────
    if prev_snap is not None and (added_wells or removed_wells):
        wc = wb.create_sheet("Well Changes")
        wc.sheet_view.showGridLines = False
        wc.freeze_panes = "A2"
        prev_ts = prev_snap.get("run_timestamp", "unknown")
        wc.cell(1, 1, f"Changes vs run at {prev_ts}").font = ITALIC
        wc.merge_cells("A1:B1")
        _hdr(wc, 2, ["Well Name", "Change"], [30, 38])
        wc.auto_filter.ref = f"A2:B{2 + len(added_wells) + len(removed_wells)}"
        ri = 3
        for name in added_wells:
            wc.cell(ri, 1, name).fill = ADDED_FILL
            wc.cell(ri, 2, "Now included  (was excluded)").fill = ADDED_FILL
            ri += 1
        for name in removed_wells:
            wc.cell(ri, 1, name).fill = EXCL_FILL
            wc.cell(ri, 2, "Now excluded  (was included)").fill = EXCL_FILL
            ri += 1

    # ── Sheet 5: Chemistry by Well ────────────────────────────────────────────
    if per_well_chem is not None and not per_well_chem.empty:
        _data_sheet(
            "Chem by Well",
            per_well_chem,
            ["NAME", "Total Rows", "Removed (REVIEWQ)", "Removed (PURPOSE)", "Remaining", "% Retained"],
            CHEM_FILL,
        )

    wb.save(out_path)
    print(f"Excel report  → {out_path}")


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> int:
    config_path = Path(sys.argv[1]) if len(sys.argv) > 1 else Path(DEFAULT_CONFIG)

    if not config_path.exists():
        print(f"ERROR: config not found: {config_path}", file=sys.stderr)
        return 2

    with open(config_path, "rb") as fh:
        raw = tomllib.load(fh)

    # ── Well filters ──────────────────────────────────────────────────────────
    try:
        well_file, cols, modes, values = _load_well_config(raw)
    except Exception as exc:
        print(f"ERROR reading well filter config: {exc}", file=sys.stderr)
        return 2

    if not well_file.exists():
        print(f"ERROR: well file not found: {well_file}", file=sys.stderr)
        return 2

    if cols and not (len(cols) == len(modes) == len(values)):
        print("ERROR: well_filter_cols, well_filter_modes, and well_filter_values must have the same length.",
              file=sys.stderr)
        return 2

    from tta.utils import load_table
    try:
        well = load_table(well_file)
    except Exception as exc:
        print(f"ERROR loading well file: {exc}", file=sys.stderr)
        return 2

    if "NAME" not in well.columns:
        print("ERROR: well file has no 'NAME' column.", file=sys.stderr)
        return 2

    missing_cols = [c for c in cols if c not in well.columns]
    if missing_cols:
        print(f"ERROR: filter column(s) not found in well file: {missing_cols}", file=sys.stderr)
        print(f"  Available: {sorted(well.columns.tolist())}", file=sys.stderr)
        return 2

    try:
        included, excluded = _apply_well_filters_stepwise(well, cols, modes, values)
    except Exception as exc:
        print(f"ERROR applying well filters: {exc}", file=sys.stderr)
        return 2

    current_included = sorted(included["NAME"].astype(str).tolist())
    n_total          = len(well)

    # ── Chemistry filters ─────────────────────────────────────────────────────
    chem_stats: Optional[dict]         = None
    per_well_chem: Optional[pd.DataFrame] = None

    try:
        chem_files, reviewq_patterns, purpose_exclude, mdl_sub, dup_handling, min_date, max_date = \
            _load_chem_config(raw)
    except Exception as exc:
        print(f"WARNING: could not read chemistry config: {exc}", file=sys.stderr)
        chem_files, reviewq_patterns, purpose_exclude, mdl_sub, dup_handling = [], [], [], 1.0, "daily_avg"
        min_date = max_date = pd.Timestamp("2000-01-01")

    chem_files_missing = [f for f in chem_files if not f.exists()]
    if chem_files_missing:
        print(f"NOTE: chemistry file(s) not found — chemistry validation skipped:")
        for f in chem_files_missing:
            print(f"  {f}")
    else:
        try:
            chem_stats, per_well_chem = _apply_chem_filters_stepwise(
                chem_files, reviewq_patterns, purpose_exclude, min_date, max_date
            )
        except Exception as exc:
            print(f"WARNING: chemistry filter validation failed: {exc}", file=sys.stderr)

    # ── Snapshot and diff ─────────────────────────────────────────────────────
    prev_snap    = _load_snapshot()
    added_wells, removed_wells = [], []
    chem_diff: dict = {}

    if prev_snap is not None:
        added_wells, removed_wells = _compute_well_diff(prev_snap, current_included)
        if chem_stats:
            chem_diff = _compute_chem_diff(prev_snap, chem_stats)

    # ── Console output ────────────────────────────────────────────────────────
    print(f"\nWELL FILTERS")
    print(f"  Well file : {well_file}  ({n_total} total)")
    print(f"  Included  : {len(included)}")
    print(f"  Excluded  : {len(excluded)}")
    if not excluded.empty and "EXCLUSION_REASON" in excluded.columns:
        for reason, count in excluded["EXCLUSION_REASON"].value_counts().items():
            print(f"    {int(count):>4}  removed by  {reason}")

    if prev_snap:
        prev_ts = prev_snap.get("run_timestamp", "unknown")
        print(f"\n  Vs previous run ({prev_ts}):")
        if added_wells or removed_wells:
            print(f"    +{len(added_wells)} now included,  -{len(removed_wells)} now excluded")
            if added_wells:
                s = ", ".join(added_wells[:8]) + (f"  … +{len(added_wells)-8} more" if len(added_wells) > 8 else "")
                print(f"    Added   : {s}")
            if removed_wells:
                s = ", ".join(removed_wells[:8]) + (f"  … +{len(removed_wells)-8} more" if len(removed_wells) > 8 else "")
                print(f"    Removed : {s}")
        else:
            print("    No change in included wells")
    else:
        print("\n  (No previous snapshot — run again after editing the config to see a diff)")

    if chem_stats:
        print(f"\nCHEMISTRY FILTERS")
        print(f"  Rows in file              : {chem_stats['n_raw']:,}")
        print(f"  After date range          : {chem_stats['n_after_date']:,}  ({chem_stats['date_range']})")
        print(f"  Removed by REVIEWQ        : {chem_stats['n_removed_reviewq']:,}")
        if chem_stats["reviewq_breakdown"]:
            for pat, count in chem_stats["reviewq_breakdown"].items():
                print(f"    pattern '{pat}': {count:,} rows")
        print(f"  Removed by PURPOSE        : {chem_stats['n_removed_purpose']:,}")
        if chem_stats["purpose_breakdown"]:
            for code, count in sorted(chem_stats["purpose_breakdown"].items(), key=lambda x: -x[1]):
                print(f"    {code}: {count:,} rows")
        print(f"  Remaining rows            : {chem_stats['n_final']:,}")
        print(f"  Wells with rows removed   : {chem_stats['n_wells_affected']}")
        if chem_stats["n_wells_emptied"]:
            print(f"  Wells completely emptied  : {chem_stats['n_wells_emptied']}  *** WARNING ***")
        if chem_diff:
            print(f"\n  Vs previous run ({prev_snap.get('run_timestamp', 'unknown')}):")
            d = chem_diff
            print(f"    REVIEWQ removed rows   : {'+' if d['d_removed_reviewq']>=0 else ''}{d['d_removed_reviewq']:,}")
            print(f"    PURPOSE removed rows   : {'+' if d['d_removed_purpose']>=0 else ''}{d['d_removed_purpose']:,}")
            print(f"    Remaining rows         : {'+' if d['d_final']>=0 else ''}{d['d_final']:,}")

    print()

    # ── Write Excel ───────────────────────────────────────────────────────────
    _write_excel(
        config_path=config_path,
        well_file=well_file,
        cols=cols, modes=modes, values=values,
        included=included, excluded=excluded,
        prev_snap=prev_snap,
        added_wells=added_wells, removed_wells=removed_wells,
        chem_files=chem_files,
        reviewq_patterns=reviewq_patterns,
        purpose_exclude=purpose_exclude,
        mdl_sub=mdl_sub,
        duplicate_handling=dup_handling,
        chem_stats=chem_stats,
        per_well_chem=per_well_chem,
        chem_diff=chem_diff,
        out_path=Path(OUTPUT_XLSX),
    )

    # ── Save snapshot ─────────────────────────────────────────────────────────
    _save_snapshot(well_file, cols, modes, values, n_total, current_included, chem_stats)
    print(f"Snapshot   → {SNAPSHOT_FILE}")

    return 0


if __name__ == "__main__":
    sys.exit(main())
